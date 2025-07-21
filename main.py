
from analysis import run_analysis
from highlight import highlight_side_lane
from delivery_helpers import generate_times

import pandas as pd
import math
from datetime import datetime
from openpyxl import load_workbook
import numpy as np
import io
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import PatternFill

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select BOM Excel File",
        filetypes=[("Excel files", "*.xlsx *.xlsm")]
    )

    print("Available Drive Units: Proteus, Hercules, Megasus")
    input_units = input("Enter Drive Units to plan (comma-separated, or 'all' for all units): ").strip().lower()
    valid_units = ['Proteus', 'Hercules', 'Megasus']

    if input_units == 'all':
        units_to_plan = valid_units
    else:
        units_to_plan = [u.strip().capitalize() for u in input_units.split(",")]
        for u in units_to_plan:
            if u not in valid_units:
                raise ValueError(f"Invalid drive unit: {u}")

    df_bom_init = pd.read_excel(file_path, sheet_name=f"Inbound-{valid_units[0]}", skiprows=15, header=None)
    df_bom_init.columns = ['Part Number', 'Description', 'Quantity / Unit', 'Needed per day',
        'Quantity Needed for Shift 1', 'Quantity Needed for Shift 2',
        'Pallets Utilized for Shift 1', 'Pallets Utilized for Shift 2',
        'Consumption Rate Units/ Hour Shift 1', 'Consumption Rate / Hour Shift 2',
        'Standard Pack Size', 'Package Type', 'Maximum Storage on Lineside',
        'Minimum Storage on Lineside', 'On-hand qty', 'QTY vs Shift 1', 'On-hand on dock', 'On-hand QTY at Lineside']
    
    #Creating dictionary for inventory on hand across all units
    inventory_on_hand = {
        part: qty for part, qty in zip(df_bom_init['Part Number'], df_bom_init['On-hand qty'])
    }
    dock_on_hand = { 
        part:qty for part, qty in zip(df_bom_init['Part Number'], df_bom_init['On-hand on dock'])
    }
    # Determine Max Cadence to align deliveries per unit 
    cadence_shift_1_list = []
    cadence_shift_2_list = []
    for unit in units_to_plan:
        wb = load_workbook(file_path, data_only=True)
        ws = wb[f'Inbound-{unit}']
        c1 = ws['B5'].value
        c2 = ws['B13'].value

        c1 = int(c1) if c1 is not None and str(c1).isdigit() else 0
        c2 = int(c2) if c2 is not None and str(c2).isdigit() else 0

        cadence_shift_1_list.append(c1)
        cadence_shift_2_list.append(c2)

    max_cadence_1 = max(cadence_shift_1_list)
    max_cadence_2 = max(cadence_shift_2_list)

    # Generate aligned delivery slots
    start_shift_1 = datetime.strptime("6:15", "%H:%M")
    end_shift_1 = datetime.strptime("15:00", "%H:%M")
    start_shift_2 = datetime.strptime("15:00", "%H:%M")
    end_shift_2 = datetime.strptime("23:15", "%H:%M")

    time_1 = generate_times(start_shift_1, end_shift_1, max_cadence_1)
    time_2 = generate_times(start_shift_2, end_shift_2, max_cadence_2)


    combined_filename = "Delivery_Inventory_Plan_All.xlsx"
    unit_lanes = {}

    with pd.ExcelWriter(combined_filename, engine='openpyxl') as writer:
        for unit in units_to_plan:
            print(f"Running analysis for {unit}...")
            df_output, df_dock_space, side_lane, lane = run_analysis(file_path, unit, inventory_on_hand, dock_on_hand, time_1, time_2)
            unit_lanes[unit] = (side_lane, lane)

            df_output.to_excel(writer, sheet_name=f"{unit}-Delivery", index=False)
            df_dock_space.to_excel(writer, sheet_name=f"{unit}-DockSpace", index=False)

    for unit in units_to_plan:
        side_lane, lane = unit_lanes[unit]
        highlight_side_lane(combined_filename, f"{unit}-Delivery", side_lane, lane)
        highlight_side_lane(combined_filename, f"{unit}-DockSpace", side_lane, lane)


trailer_capacity_pallets = 96  # per slot

print(f"Balancing deliveries per part/unit with trailer capacity = {trailer_capacity_pallets} boxes per slot…")

# === Read all delivery sheets and add Drive Unit column ===
delivery_dfs = []
for unit in units_to_plan:
    df = pd.read_excel(combined_filename, sheet_name=f"{unit}-Delivery")
    df['Drive Unit'] = unit  # add unit info
    delivery_dfs.append(df)

combined_deliveries = pd.concat(delivery_dfs, ignore_index=True)

# Identify delivery columns
delivery_cols = [col for col in combined_deliveries.columns if 'Delivery' in col] 

# Convet delivery units to pack size 
for col in delivery_cols:
    combined_deliveries[col] = np.ceil(combined_deliveries[col] / combined_deliveries['Pack Size'])
    is_box = combined_deliveries['Package Type'].str.lower() == 'box'
    combined_deliveries.loc[is_box, col] = np.ceil(combined_deliveries.loc[is_box, col] / 8)

slot_totals = pd.Series(0, index=delivery_cols)
balanced_df = combined_deliveries.copy()

# === Balance per part/unit row ===
for idx, row in balanced_df.iterrows():
    planned = row[delivery_cols].values.astype(int)
    balanced = [0] * len(delivery_cols)

    for i, qty in enumerate(planned):
        # Available space in this slot
        available = max(trailer_capacity_pallets - slot_totals.iloc[i], 0)

        assigned = min(qty, available)
        balanced[i] = assigned
        slot_totals.iloc[i] += assigned

        overflow = qty - assigned

        # Push overflow to later slots
        j = i + 1
        while overflow > 0 and j < len(delivery_cols):
            avail_j = max(trailer_capacity_pallets - slot_totals.iloc[j], 0)
            assign_j = min(overflow, avail_j)
            balanced[j] += assign_j
            slot_totals.iloc[j] += assign_j
            overflow -= assign_j
            j += 1

        if overflow > 0:
            print(f"Part {row['Part Number']} ({row['Drive Unit']}) has {overflow} boxes unassigned after last slot.")

    # Update balanced row
    balanced_df.loc[idx, delivery_cols] = balanced

# === Add TOTAL row ===
total_row = {
    'Drive Unit': 'TOTAL',
    'Part Number': '',
    'Package Type': '',
    'Description': '',
    'Pack Size': ''
}

for col in delivery_cols:
    total_row[col] = balanced_df[col].sum()

# Append total row
balanced_df = pd.concat([
    balanced_df,
    pd.DataFrame([total_row])
], ignore_index=True)

# === Save to Excel ===
with pd.ExcelWriter(combined_filename, mode='a', engine='openpyxl') as writer:
    balanced_df = balanced_df.loc[:, ~balanced_df.columns.str.contains('^Unnamed')]
    balanced_df.to_excel(writer, sheet_name='Balanced_Part_Deliveries', index=False)


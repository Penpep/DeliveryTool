from analysis import run_analysis
from highlight import highlight_side_lane
from delivery_helpers import generate_times

import streamlit as st
import pandas as pd
import math
from datetime import datetime
from openpyxl import load_workbook
import numpy as np
from io import BytesIO
from openpyxl.styles import PatternFill



col1, col2 = st.columns([1, 6])  # adjust ratio

with col1:
    st.image("logo.png", width=100)

with col2:
    st.title("Inbound Delivery Optimization Tool")
    st.write("This tool is for optimizing move orders using BOM data and according to trailer contraints across multiple drive units")

uploaded_file = st.file_uploader("Upload BOM Excel File", type=["xlsx", "xlsm"])
if uploaded_file is None:
    st.stop()
valid_units = ['Proteus', 'Hercules', 'Megasus']
units_to_plan = st.multiselect("Select Drive Units to plan", valid_units, default=valid_units)

st.write("Max Trailer Capacity is 48 pallets")
trailer_capacity_pallets = st.number_input(
    "Trailer Capacity (pallets per slot)", min_value=1, value=144
)
st.image("Layout.png", caption="Labeled Layout of CMA Area to go along with Excel Sheet", use_container_width=True)
if st.button("Run Analysis & Balance Deliveries"):

    combined_filename = "Delivery_Inventory_Plan_All.xlsx"
    unit_lanes = {}

    # Read BOM initial sheet
    df_bom_init = pd.read_excel(uploaded_file, sheet_name=f"Inbound-{valid_units[0]}", skiprows=15, header=None)
    df_bom_init.columns = ['Part Number', 'Description', 'Quantity / Unit', 'Needed per day',
        'Quantity Needed for Shift 1', 'Quantity Needed for Shift 2',
        'Pallets Utilized for Shift 1', 'Pallets Utilized for Shift 2',
        'Consumption Rate Units/ Hour Shift 1', 'Consumption Rate / Hour Shift 2',
        'Standard Pack Size', 'Package Type', 'Maximum Storage on Lineside',
        'Minimum Storage on Lineside', 'On-hand qty', 'QTY vs Shift 1', 'On-hand on dock', 'On-hand QTY at Lineside']

    inventory_on_hand = {
        part: qty for part, qty in zip(df_bom_init['Part Number'], df_bom_init['On-hand qty'])
    }
    dock_on_hand = {
        part: qty for part, qty in zip(df_bom_init['Part Number'], df_bom_init['On-hand on dock'])
    }

    # Determine max cadences
    cadence_shift_1_list, cadence_shift_2_list = [], []
    for unit in units_to_plan:
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb[f'Inbound-{unit}']
        c1 = int(ws['B5'].value or 0)
        c2 = int(ws['B13'].value or 0)
        cadence_shift_1_list.append(c1)
        cadence_shift_2_list.append(c2)

    max_cadence_1 = max(cadence_shift_1_list)
    max_cadence_2 = max(cadence_shift_2_list)

    start_shift_1 = datetime.strptime("6:15", "%H:%M")
    end_shift_1 = datetime.strptime("15:00", "%H:%M")
    start_shift_2 = datetime.strptime("15:00", "%H:%M")
    end_shift_2 = datetime.strptime("23:15", "%H:%M")

    time_1 = generate_times(start_shift_1, end_shift_1, max_cadence_1)
    time_2 = generate_times(start_shift_2, end_shift_2, max_cadence_2)

    # === Write initial Excel to disk ===
    with pd.ExcelWriter(combined_filename, engine='openpyxl') as writer:
        for unit in units_to_plan:
            df_output, df_dock_space, side_lane, lane = run_analysis(
                uploaded_file, unit, inventory_on_hand, dock_on_hand, time_1, time_2
            )
            unit_lanes[unit] = (side_lane, lane)

            df_output.to_excel(writer, sheet_name=f"{unit}-Delivery", index=False)
            df_dock_space.to_excel(writer, sheet_name=f"{unit}-DockSpace", index=False)

    # === Apply highlights ===
    for unit in units_to_plan:
        side_lane, lane = unit_lanes[unit]
        highlight_side_lane(combined_filename, f"{unit}-Delivery", side_lane, lane)
        highlight_side_lane(combined_filename, f"{unit}-DockSpace", side_lane, lane)

    # === Add balanced deliveries ===
    delivery_dfs = []
    for unit in units_to_plan:
        df = pd.read_excel(combined_filename, sheet_name=f"{unit}-Delivery")
        df['Drive Unit'] = unit
        delivery_dfs.append(df)

    combined_deliveries = pd.concat(delivery_dfs, ignore_index=True)
    delivery_cols = [col for col in combined_deliveries.columns if 'Delivery' in col]

    for col in delivery_cols:
        combined_deliveries[col] = np.ceil(combined_deliveries[col] / combined_deliveries['Pack Size'])
        is_box = combined_deliveries['Package Type'].str.lower() == 'box'
        combined_deliveries.loc[is_box, col] = np.ceil(combined_deliveries.loc[is_box, col] / 8)

    slot_totals = pd.Series(0, index=delivery_cols)
    balanced_df = combined_deliveries.copy()

    for idx, row in balanced_df.iterrows():
        planned = row[delivery_cols].values.astype(int)
        balanced = [0] * len(delivery_cols)

        for i, qty in enumerate(planned):
            available = max(trailer_capacity_pallets - slot_totals.iloc[i], 0)
            assigned = min(qty, available)
            balanced[i] = assigned
            slot_totals.iloc[i] += assigned

            overflow = qty - assigned
            j = i + 1
            while overflow > 0 and j < len(delivery_cols):
                avail_j = max(trailer_capacity_pallets - slot_totals.iloc[j], 0)
                assign_j = min(overflow, avail_j)
                balanced[j] += assign_j
                slot_totals.iloc[j] += assign_j
                overflow -= assign_j
                j += 1

        balanced_df.loc[idx, delivery_cols] = balanced

    total_row = { 'Drive Unit': 'TOTAL', 'Part Number': '', 'Package Type': '', 'Description': '', 'Pack Size': '' }
    for col in delivery_cols:
        total_row[col] = balanced_df[col].sum()
    balanced_df = pd.concat([balanced_df, pd.DataFrame([total_row])], ignore_index=True)

    balanced_df = balanced_df.loc[:, ~balanced_df.columns.str.contains('^Unnamed')]
    with pd.ExcelWriter(combined_filename, mode='a', engine='openpyxl') as writer:
        balanced_df.to_excel(writer, sheet_name='Balanced_Part_Deliveries', index=False)

    # === Read back to memory ===
    with open(combined_filename, "rb") as f:
        output_buffer = BytesIO(f.read())

    st.success("Download your plan below.")
    st.download_button(
        label="Download Delivery Plan",
        data=output_buffer,
        file_name=combined_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
